import uuid
from django.contrib.auth import get_user_model
from django.utils import timezone
from datetime import timedelta
from rest_framework import generics, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenRefreshView as JWTTokenRefreshView

from .models import PasswordResetToken
from .permissions import IsAdminHR, IsSuperAdmin
from .serializers import (
    RegisterSerializer,
    LoginSerializer,
    UserSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    UserPasswordResetSerializer,
    ChangePasswordSerializer,
    PasswordResetRequestSerializer,
    PasswordResetConfirmSerializer,
)

User = get_user_model()

# -------------------------------------------------------
# Authentication & User Management Views
# Handles login, logout, registration, password reset
# and full user CRUD for Super Admin
# -------------------------------------------------------

class RegisterView(generics.CreateAPIView):
    """POST /api/v1/auth/register/ -- Admin-only user creation."""
    serializer_class = RegisterSerializer
    permission_classes = [IsAdminHR]


class LoginView(APIView):
    """POST /api/v1/auth/login/ -- Returns access + refresh tokens."""
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email']
        password = serializer.validated_data['password']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response(
                {'detail': 'Invalid credentials.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        if not user.check_password(password):
            return Response(
                {'detail': 'Invalid credentials.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        if not user.is_active:
            return Response(
                {'detail': 'Account is disabled.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # Generate JWT tokens for the authenticated user
        refresh = RefreshToken.for_user(user)
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'user': UserSerializer(user).data,
        })


class LogoutView(APIView):
    """POST /api/v1/auth/logout/ -- Blacklists the refresh token."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
            return Response({'detail': 'Successfully logged out.'}, status=status.HTTP_200_OK)
        except Exception:
            return Response({'detail': 'Invalid token.'}, status=status.HTTP_400_BAD_REQUEST)


# Token refresh endpoint - extends SimpleJWT default view
class TokenRefreshView(JWTTokenRefreshView):
    """POST /api/v1/auth/token/refresh/"""
    pass


class ProfileView(generics.RetrieveUpdateAPIView):
    """GET/PATCH /api/v1/auth/profile/ -- Current user's own profile."""
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user


class ChangePasswordView(APIView):
    """PUT /api/v1/auth/change-password/"""
    permission_classes = [IsAuthenticated]

    def put(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = request.user
        if not user.check_password(serializer.validated_data['old_password']):
            return Response(
                {'detail': 'Old password is incorrect.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user.set_password(serializer.validated_data['new_password'])
        user.save()
        return Response({'detail': 'Password changed successfully.'})


class PasswordResetRequestView(APIView):
    """POST /api/v1/auth/password-reset/ -- Sends email with reset token."""
    permission_classes = [AllowAny]

    def post(self, request):
        from django.conf import settings as django_settings

        serializer = PasswordResetRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email']
        response_data = {'detail': 'If the email exists, a reset link has been sent.'}

        try:
            user = User.objects.get(email=email)
            token = str(uuid.uuid4())
            PasswordResetToken.objects.create(
                user=user,
                token=token,
                expires_at=timezone.now() + timedelta(hours=24),
            )
            # In production, send email with token via configured email backend
            # In development (DEBUG=True), return the token directly in the response
            # so the feature is testable without an email server.
            if django_settings.DEBUG:
                response_data['debug_token'] = token
                response_data['debug_note'] = (
                    'Token returned only in DEBUG mode. '
                    'In production, this would be sent via email.'
                )
        except User.DoesNotExist:
            pass  # Don't reveal whether email exists

        return Response(response_data)


class PasswordResetConfirmView(APIView):
    """POST /api/v1/auth/password-reset/confirm/ -- Validates token, sets new password."""
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PasswordResetConfirmSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            reset_token = PasswordResetToken.objects.get(
                token=serializer.validated_data['token'],
                is_used=False,
                expires_at__gt=timezone.now(),
            )
        except PasswordResetToken.DoesNotExist:
            return Response(
                {'detail': 'Invalid or expired token.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = reset_token.user
        user.set_password(serializer.validated_data['new_password'])
        user.save()

        reset_token.is_used = True
        reset_token.save()

        return Response({'detail': 'Password has been reset successfully.'})


class UserListView(generics.ListAPIView):
    """GET /api/v1/auth/users/ -- Super Admin only user listing."""
    queryset = User.objects.all().order_by('-created_at')
    serializer_class = UserSerializer
    permission_classes = [IsSuperAdmin]
    search_fields = ['first_name', 'last_name', 'email', 'username']
    filterset_fields = ['role', 'is_active']


class UserCreateView(generics.CreateAPIView):
    """POST /api/v1/auth/users/ -- Super Admin creates new user."""
    serializer_class = UserCreateSerializer
    permission_classes = [IsSuperAdmin]


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    """GET/PATCH/DELETE /api/v1/auth/users/<uuid>/ -- Super Admin user management."""
    queryset = User.objects.all()
    permission_classes = [IsSuperAdmin]
    
    def get_serializer_class(self):
        if self.request.method in ('PUT', 'PATCH'):
            return UserUpdateSerializer
        return UserSerializer


class UserPasswordResetView(APIView):
    """POST /api/v1/auth/users/<uuid>/reset-password/ -- Super Admin resets user password."""
    permission_classes = [IsSuperAdmin]
    
    def post(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({'detail': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = UserPasswordResetSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user.set_password(serializer.validated_data['new_password'])
        user.save()
        
        return Response({'detail': 'Password reset successfully.'})


class UserBulkImportView(APIView):
    """POST /api/v1/auth/users/bulk-import/ -- Super Admin imports users from Excel."""
    permission_classes = [IsSuperAdmin]
    parser_classes_list = None  # will use default

    def post(self, request):
        import openpyxl
        from io import BytesIO

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
        except Exception:
            return Response({'detail': 'Invalid Excel file.'}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
        required = {'first_name', 'last_name', 'email', 'username', 'password'}
        if not required.issubset(set(headers)):
            return Response(
                {'detail': f'Missing columns. Required: {", ".join(required)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        created, errors = [], []
        # Process each row starting from row 2 (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)}
            if not data.get('email') or not data.get('username'):
                continue
            try:
                if User.objects.filter(email=data['email']).exists():
                    errors.append({'row': row_idx, 'email': data['email'], 'error': 'Email already exists'})
                    continue
                if User.objects.filter(username=data['username']).exists():
                    errors.append({'row': row_idx, 'email': data['email'], 'error': 'Username already exists'})
                    continue
                user = User(
                    first_name=data.get('first_name', ''),
                    last_name=data.get('last_name', ''),
                    email=data['email'],
                    username=data['username'],
                    role=data.get('role', 'STUDENT').upper(),
                    phone=data.get('phone', ''),
                    is_active=True,
                )
                user.set_password(data.get('password', 'changeme123'))
                user.save()
                created.append(data['email'])
            except Exception as e:
                errors.append({'row': row_idx, 'email': data.get('email', ''), 'error': str(e)})

        return Response({
            'created': len(created),
            'errors': errors,
            'detail': f'{len(created)} user(s) created, {len(errors)} error(s).'
        })
